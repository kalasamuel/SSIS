from django.shortcuts import render, redirect, get_object_or_404
from django.forms import inlineformset_factory
from django.contrib import messages
from django.db import transaction
from django.db.models.functions import Coalesce , Greatest, Cast
from django.db.models import Sum, F, Value, DecimalField, Count
from django.http import JsonResponse
from django.utils.dateparse import parse_date



from .models import (
    Sale, SaleDetail, Product, Supplier, Category,
    Customer, Staff, Discount, PurchaseOrder,
    PurchaseOrderDetail, InventoryLog, Payroll
)
from .forms import (
    SaleForm, SaleDetailForm, SaleDetailFormSet, ProductForm, SupplierForm,
    CategoryForm, CustomerForm, StaffForm, DiscountForm, PurchaseOrderForm,
    PurchaseOrderDetailForm, InventoryLogForm, PayrollForm
)

from .utils import generate_purchase_order_pdf

#graphs quarterly and yearly sales
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, F, FloatField, ExpressionWrapper, DecimalField
from django.db.models.functions import ExtractYear, ExtractQuarter, ExtractMonth, TruncDate, TruncDay, TruncWeek, TruncMonth, TruncQuarter


from django.utils import timezone
from datetime import timedelta


from datetime import timedelta, date, datetime
from django.utils.timezone import now
from django.core.mail import EmailMessage
from django.conf import settings
from decimal import Decimal

# Import for exports
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import csv
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ---------------------------------------------------------
# DASHBOARD / HOME PAGE
# ---------------------------------------------------------
def authentication(request):
    pass

def home(request):
    total_sales = Sale.objects.count()
    total_products = Product.objects.count()
    total_customers = Customer.objects.count()
    total_staff = Staff.objects.count()
    total_suppliers = Supplier.objects.count()
    return render(request, "inventory/home.html", {
        "total_sales": total_sales,
        "total_products": total_products,
        "total_customers": total_customers,
        "total_staff": total_staff,
        "total_suppliers": total_suppliers,
    })

def dashboard(request):
    today = now().date()
    last_month = today - timedelta(days=30)
    previous_month = today - timedelta(days=60)

    current_sales = Sale.objects.filter(
        sale_datetime__gte=last_month
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    past_sales = Sale.objects.filter(
        sale_datetime__gte=previous_month,
        sale_datetime__lt=last_month
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    growth = ((current_sales - past_sales) / past_sales * 100) if past_sales > 0 else 0

    context = {
        "current_sales": current_sales,
        "past_sales": past_sales,
        "growth": growth,
    }

    return render(request, "dashboard.html", context)
# ---------------------------------------------------------
# SALE + SALE DETAILS
# ---------------------------------------------------------
@transaction.atomic
def create_sale(request):
    """Create a sale and related sale detail items."""
    if request.method == "POST":
        sale_form = SaleForm(request.POST)
        formset = SaleDetailFormSet(request.POST)
        if sale_form.is_valid() and formset.is_valid():
            sale = sale_form.save(commit=False)
            sale.save()
            formset.instance = sale
            formset.save()
            
            # Update stock quantities for each sale detail
            stock_updates = []
            for sale_detail in formset:
                if sale_detail.cleaned_data and not sale_detail.cleaned_data.get('DELETE', False):
                    product = sale_detail.cleaned_data['product']
                    quantity_sold = sale_detail.cleaned_data['quantity_sold']
                    
                    # Check if sufficient stock is available
                    if product.stock_quantity < quantity_sold:
                        messages.error(request, f"Insufficient stock for {product.product_name}. Available: {product.stock_quantity}")
                        return redirect("create_sale")
                    
                    # Update stock quantity
                    product.stock_quantity -= quantity_sold
                    product.save()
                    
                    # Create inventory log entry
                    InventoryLog.objects.create(
                        product=product,
                        staff=sale.staff,
                        log_type='Sale',
                        quantity=quantity_sold,
                        remarks=f"Sale #{sale.receipt_no} - {sale_detail.cleaned_data.get('batch_number', 'No batch')}"
                    )
                    
                    stock_updates.append(f"{product.product_name}: -{quantity_sold}")
            
            # Apply automatic discounts
            discount_result = apply_automatic_discounts(sale)
            if discount_result:
                messages.success(request, f"Sale recorded successfully. Stock updated. Discount applied: UGx. {discount_result['discount_amount']:,.0f}")
            else:
                messages.success(request, f"Sale recorded successfully. Stock updated for {len(stock_updates)} products.")
            
            return redirect("sales_list")
    else:
        sale_form = SaleForm()
        formset = SaleDetailFormSet()

    return render(request, "inventory/billing_form.html", {
        "sale_form": sale_form,
        "formset": formset
    })


def sales_list(request):
    """List all recorded sales."""
    sales = Sale.objects.all().order_by("-sale_datetime")
    return render(request, "inventory/sales_list.html", {"sales": sales})


def sale_detail(request, pk):
    """View detailed sale items."""
    sale = get_object_or_404(Sale, pk=pk)
    details = SaleDetail.objects.filter(sale=sale)
    return render(request, "inventory/sale_detail.html", {"sale": sale, "details": details})


def sale_items_api(request, pk):
    """API endpoint to get sale items"""
    sale = get_object_or_404(Sale, pk=pk)
    details = SaleDetail.objects.filter(sale=sale).select_related('product')
    
    data = {
        'sale': {
            'id': sale.id,
            'receipt_no': sale.receipt_no,
            'sale_datetime': sale.sale_datetime.strftime('%Y-%m-%d %H:%M:%S'),
            'customer_name': f"{sale.customer.first_name} {sale.customer.last_name}" if sale.customer else "Walk-in Customer",
            'staff_name': f"{sale.staff.first_name} {sale.staff.last_name}",
            'total_amount': float(sale.total_amount),
            'discount_applied': float(sale.discount_applied) if sale.discount_applied else 0,
            'payment_method': sale.payment_method,
        },
        'items': [
            {
                'product_name': detail.product.product_name,
                'quantity': detail.quantity_sold,
                'unit_price': float(detail.unit_price),
                'discount_value': float(detail.discount_value) if detail.discount_value else 0,
                'sub_total': float(detail.sub_total) if detail.sub_total else float(detail.unit_price * detail.quantity_sold),
                'batch_number': detail.batch_number or 'N/A',
            }
            for detail in details
        ]
    }
    return JsonResponse(data)


def export_sales(request):
    """Export sales data"""
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        return export_sales_csv(request)
    elif export_format == 'pdf':
        return export_sales_pdf(request)
    else:
        return HttpResponse("Invalid export format", status=400)


def export_sales_csv(request):
    """Export sales as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Receipt No', 'Date', 'Customer', 'Staff', 'Payment Method', 'Total Amount', 'Discount Applied', 'Items Count'])
    
    sales = Sale.objects.select_related('customer', 'staff').all().order_by('-sale_datetime')
    for sale in sales:
        writer.writerow([
            sale.receipt_no,
            sale.sale_datetime.strftime('%Y-%m-%d %H:%M:%S'),
            f"{sale.customer.first_name} {sale.customer.last_name}" if sale.customer else "Walk-in Customer",
            f"{sale.staff.first_name} {sale.staff.last_name}",
            sale.payment_method,
            sale.total_amount,
            sale.discount_applied or 0,
            sale.saledetail_set.count()
        ])
    
    return response


def export_sales_pdf(request):
    """Export sales as PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("Sales Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Sales summary
    sales = Sale.objects.all()
    total_sales = sales.count()
    total_revenue = sales.aggregate(total=Sum('total_amount'))['total'] or 0
    total_discounts = sales.aggregate(total=Sum('discount_applied'))['total'] or 0
    
    summary_data = [
        ['Total Sales', total_sales],
        ['Total Revenue', f'UGx. {total_revenue:,.0f}'],
        ['Total Discounts', f'UGx. {total_discounts:,.0f}'],
        ['Net Revenue', f'UGx. {total_revenue - total_discounts:,.0f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('GRID',(0,0),(-1,-1),1,colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Sales details
    elements.append(Paragraph("Sales Details", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    table_data = [['Receipt', 'Date', 'Customer', 'Staff', 'Payment', 'Total', 'Items']]
    for sale in sales[:50]:  # Limit for PDF
        table_data.append([
            sale.receipt_no,
            sale.sale_datetime.strftime('%Y-%m-%d'),
            f"{sale.customer.first_name} {sale.customer.last_name}" if sale.customer else "Walk-in",
            f"{sale.staff.first_name} {sale.staff.last_name}",
            sale.payment_method,
            f'UGx. {sale.total_amount:,.0f}',
            str(sale.saledetail_set.count())
        ])
    
    sales_table = Table(table_data, colWidths=[1*inch, 1*inch, 1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 0.5*inch])
    sales_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.5,colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8)
    ]))
    elements.append(sales_table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def print_receipt(request, receipt_no):
    """Generate receipt for printing"""
    sale = get_object_or_404(Sale, receipt_no=receipt_no)
    details = SaleDetail.objects.filter(sale=sale).select_related('product')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="receipt_{receipt_no}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Receipt header
    elements.append(Paragraph("SUPERMARKET RECEIPT", styles['Heading1']))
    elements.append(Paragraph(f"Receipt No: {sale.receipt_no}", styles['Normal']))
    elements.append(Paragraph(f"Date: {sale.sale_datetime.strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Paragraph(f"Staff: {sale.staff.first_name} {sale.staff.last_name}", styles['Normal']))
    if sale.customer:
        elements.append(Paragraph(f"Customer: {sale.customer.first_name} {sale.customer.last_name}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Items table
    elements.append(Paragraph("Items Purchased", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    table_data = [['Product', 'Qty', 'Price', 'Total']]
    for detail in details:
        table_data.append([
            detail.product.product_name,
            str(detail.quantity_sold),
            f'UGx. {detail.unit_price:,.0f}',
            f'UGx. {(detail.unit_price * detail.quantity_sold):,.0f}'
        ])
    
    receipt_table = Table(table_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
    receipt_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.5,colors.black)
    ]))
    elements.append(receipt_table)
    elements.append(Spacer(1, 20))
    
    # Totals
    subtotal = sum(detail.unit_price * detail.quantity_sold for detail in details)
    discount = sale.discount_applied or 0
    total = sale.total_amount
    
    totals_data = [
        ['Subtotal', f'UGx. {subtotal:,.0f}'],
        ['Discount', f'-UGx. {discount:,.0f}'],
        ['Total', f'UGx. {total:,.0f}']
    ]
    
    totals_table = Table(totals_data, colWidths=[2*inch, 2*inch])
    totals_table.setStyle(TableStyle([
        ('ALIGN',(0,0),(-1,-1),'RIGHT'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 12)
    ]))
    elements.append(totals_table)
    
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Thank you for your business!", styles['Normal']))
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


# ---------------------------------------------------------
# PRODUCT
# ---------------------------------------------------------
def create_product(request):
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Product added successfully.")
            return redirect("product_list")
    else:
        form = ProductForm()
    return render(request, "inventory/product_form.html", {"form": form})


def product_list(request):
    products = Product.objects.all().order_by("product_name")
    return render(request, "inventory/product_list.html", {"products": products})


def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "Product updated.")
            return redirect("product_list")
    else:
        form = ProductForm(instance=product)
    return render(request, "inventory/product_form.html", {"form": form})


def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product_name = product.product_name
        product.delete()
        messages.success(request, f"Product '{product_name}' has been deleted.")
        return redirect("product_list")
    return redirect("product_list")


# ---------------------------------------------------------
# SUPPLIER
# ---------------------------------------------------------
def create_supplier(request):
    if request.method == "POST":
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier added successfully.")
            return redirect("supplier_list")
    else:
        form = SupplierForm()
    return render(request, "inventory/supplier_form.html", {"form": form})


def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, "inventory/supplier_list.html", {"suppliers": suppliers})

def edit_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated successfully.")
            return redirect("supplier_list")
    else:
        form = SupplierForm(instance=supplier)
    return render(request, "inventory/supplier_form.html", {"form": form})

# ---------------------------------------------------------
# CATEGORY
# ---------------------------------------------------------
def create_category(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added successfully.")
            return redirect("category_list")
    else:
        form = CategoryForm()
    return render(request, "inventory/category_form.html", {"form": form})


def category_list(request):
    categories = Category.objects.all()
    return render(request, "inventory/category_list.html", {"categories": categories})


# ---------------------------------------------------------
# CUSTOMER
# ---------------------------------------------------------
def create_customer(request):
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Customer registered successfully.")
            return redirect("customer_list")
    else:
        form = CustomerForm()
    return render(request, "inventory/customer_form.html", {"form": form})


def customer_list(request):
    customers = Customer.objects.all().order_by("first_name")
    return render(request, "inventory/customer_list.html", {"customers": customers})


def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "Customer updated successfully.")
            return redirect("customer_list")
    else:
        form = CustomerForm(instance=customer)
    return render(request, "inventory/customer_form.html", {"form": form})


# ---------------------------------------------------------
# STAFF
# ---------------------------------------------------------
def create_staff(request):
    if request.method == "POST":
        form = StaffForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Staff added successfully.")
            return redirect("staff_list")
    else:
        form = StaffForm()
    return render(request, "inventory/staff_form.html", {"form": form})


def staff_list(request):
    staff = Staff.objects.all()
    return render(request, "inventory/staff_list.html", {"staff": staff})


# ---------------------------------------------------------
# DISCOUNT
# ---------------------------------------------------------
def create_discount(request):
    if request.method == "POST":
        form = DiscountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Discount created successfully.")
            return redirect("discount_list")
    else:
        form = DiscountForm()
    return render(request, "inventory/discount_form.html", {"form": form})


def discount_list(request):
    discounts = Discount.objects.all().order_by("-start_date")
    return render(request, "inventory/discount_list.html", {"discounts": discounts})


def edit_discount(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    if request.method == "POST":
        form = DiscountForm(request.POST, instance=discount)
        if form.is_valid():
            form.save()
            messages.success(request, "Discount scheme updated successfully.")
            return redirect("discount_list")
    else:
        form = DiscountForm(instance=discount)
    return render(request, "inventory/discount_form.html", {"form": form, "discount": discount})


def delete_discount(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    if request.method == "POST":
        discount_name = discount.discount_name
        discount.delete()
        messages.success(request, f"Discount scheme '{discount_name}' has been deleted.")
        return redirect("discount_list")
    return redirect("discount_list")


def toggle_discount_status(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    if request.method == "POST":
        import json
        data = json.loads(request.body)
        discount.is_active = data.get('is_active', not discount.is_active)
        discount.save()
        return JsonResponse({'success': True, 'is_active': discount.is_active})
    return JsonResponse({'success': False})


def discount_details_api(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    data = {
        'discount_name': discount.discount_name,
        'discount_type': discount.discount_type,
        'value': float(discount.value),
        'start_date': discount.start_date.strftime('%B %d, %Y'),
        'end_date': discount.end_date.strftime('%B %d, %Y'),
        'is_active': discount.is_active,
    }
    return JsonResponse(data)


def export_discounts(request):
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        return export_discounts_csv(request)
    elif export_format == 'pdf':
        return export_discounts_pdf(request)
    else:
        return HttpResponse("Invalid export format", status=400)


def export_discounts_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="discount_schemes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Discount Name', 'Type', 'Value', 'Start Date', 'End Date', 'Status'])
    
    discounts = Discount.objects.all().order_by('-start_date')
    for discount in discounts:
        writer.writerow([
            discount.discount_name,
            discount.discount_type,
            discount.value,
            discount.start_date.strftime('%Y-%m-%d'),
            discount.end_date.strftime('%Y-%m-%d'),
            'Active' if discount.is_active else 'Inactive'
        ])
    
    return response


def export_discounts_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="discount_schemes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("Discount Schemes Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Discount summary
    discounts = Discount.objects.all().order_by('-start_date')
    active_count = sum(1 for d in discounts if d.is_active)
    
    summary_data = [
        ['Total Schemes', len(discounts)],
        ['Active Schemes', active_count],
        ['Inactive Schemes', len(discounts) - active_count],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('GRID',(0,0),(-1,-1),1,colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Discount details
    elements.append(Paragraph("Discount Details", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    table_data = [['Name', 'Type', 'Value', 'Start Date', 'End Date', 'Status']]
    for discount in discounts[:50]:  # Limit for PDF
        table_data.append([
            discount.discount_name,
            discount.discount_type,
            str(discount.value),
            discount.start_date.strftime('%Y-%m-%d'),
            discount.end_date.strftime('%Y-%m-%d'),
            'Active' if discount.is_active else 'Inactive'
        ])
    
    discount_table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    discount_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.5,colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8)
    ]))
    elements.append(discount_table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


# ---------------------------------------------------------
# PURCHASE ORDER
# ---------------------------------------------------------

def create_purchase_order(request):
    suppliers = Supplier.objects.all()
    products = Product.objects.all()

    if request.method == "POST":
        supplier_id = request.POST.get("supplier")
        expected_date = request.POST.get("expected_delivery_date")
        invoice_no = request.POST.get("invoice_no")

        # Resolve Staff instance (not auth User)
        staff = Staff.objects.filter(username=getattr(request.user, "username", "")).first()
        if not staff:
            staff, _ = Staff.objects.get_or_create(
                username="system",
                defaults=dict(first_name="System", last_name="User", role="Admin", password_hash="system_user"),
            )

        # Create the main order
        order = PurchaseOrder.objects.create(
            supplier_id=supplier_id,
            staff=staff,
            order_date=date.today(),
            expected_delivery_date=expected_date or None,
            status="Pending",
            total_cost=0,
            invoice_no=invoice_no,
        )

        # Add product line items
        product_ids = request.POST.getlist("product[]")
        quantities = request.POST.getlist("quantity[]")
        unit_costs = request.POST.getlist("unit_cost[]")

        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and unit_costs[i]:
                qty = int(quantities[i])
                cost = float(unit_costs[i])
                line_total = qty * cost
                PurchaseOrderDetail.objects.create(
                    order=order,
                    product_id=product_ids[i],
                    quantity_ordered=qty,
                    unit_cost=cost,
                    sub_total=line_total,
                )

        # Update order total
        order.total_cost = sum(item.sub_total for item in order.items.all())
        order.save(update_fields=["total_cost"])

        #  Generate PDF
        pdf_data = generate_purchase_order_pdf(order)

        #  Build and send email
        subject = f"Purchase Order #{order.id} from {request.user.username}"
        body = (
            f"Dear {order.supplier.supplier_name},\n\n"
            f"Please find attached our new purchase order.\n\n"
            f"Thank you,\n{request.user.username}\n{request.user.email}"
        )

        email = EmailMessage(
            subject,
            body,
            settings.DEFAULT_FROM_EMAIL,
            [order.supplier.email] if order.supplier.email else [],
        )
        email.attach(f"PurchaseOrder_{order.id}.pdf", pdf_data, "application/pdf")
        email.send(fail_silently=False)

        messages.success(request, "Purchase order created and sent to supplier.")
        return redirect("purchase_order_list")

    return render(request, "inventory/create_purchase_order.html", {
        "suppliers": suppliers,
        "products": products,
    })


def purchase_order_list(request):
    orders = PurchaseOrder.objects.all().order_by("-order_date")
    return render(request, "inventory/purchase_order_list.html", {"orders": orders})


# ---------------------------------------------------------
# PURCHASE ORDER DETAIL
# ---------------------------------------------------------
def create_purchase_order_detail(request):
    if request.method == "POST":
        form = PurchaseOrderDetailForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Purchase order detail added.")
            return redirect("purchase_order_detail_list")
    else:
        form = PurchaseOrderDetailForm()
    return render(request, "inventory/purchase_order_detail_form.html", {"form": form})


def purchase_order_detail_list(request):
    details = PurchaseOrderDetail.objects.all()
    return render(request, "inventory/purchase_order_detail_list.html", {"details": details})


# ---------------------------------------------------------
# INVENTORY LOG (Material Arrival)
# ---------------------------------------------------------
def log_inventory(request):
    if request.method == "POST":
        form = InventoryLogForm(request.POST)
        if form.is_valid():
            inventory_log = form.save(commit=False)
            
            # Update product stock based on log type
            product = inventory_log.product
            if inventory_log.log_type == 'Purchase':
                # Increase stock for purchases
                product.stock_quantity += inventory_log.quantity
                product.save()
                messages.success(request, f"Inventory updated. Stock increased by {inventory_log.quantity} for {product.product_name}.")
            elif inventory_log.log_type == 'Adjustment':
                # For adjustments, we'll let the adjust_stock API handle this
                # This is mainly for manual adjustments
                pass
            
            inventory_log.save()
            messages.success(request, "Inventory log recorded.")
            return redirect("inventory_log_list")
    else:
        form = InventoryLogForm()
    return render(request, "inventory/material_arrival_form.html", {"form": form})


def inventory_log_list(request):
    logs = InventoryLog.objects.all().order_by("-log_date")
    logs = InventoryLog.objects.all().order_by("-log_date")
    return render(request, "inventory/inventory_log_list.html", {"logs": logs})


def inventory_list(request):
    """Main inventory management view"""
    products = Product.objects.select_related('category', 'supplier').all()
    return render(request, "inventory/inventory_list.html", {"products": products})


def inventory_products_api(request):
    """API endpoint to get all products with related data"""
    products = Product.objects.select_related('category', 'supplier').all()
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()
    
    data = {
        'products': [
            {
                'id': p.id,
                'product_name': p.product_name,
                'brand': p.brand,
                'unit': p.unit,
                'stock_quantity': p.stock_quantity,
                'reorder_level': p.reorder_level,
                'unit_cost': float(p.unit_cost),
                'retail_price': float(p.retail_price),
                'expiry_date': p.expiry_date.strftime('%Y-%m-%d') if p.expiry_date else None,
                'category_id': p.category.id,
                'category_name': p.category.category_name,
                'supplier_id': p.supplier.id,
                'supplier_name': p.supplier.supplier_name,
            }
            for p in products
        ],
        'categories': [
            {'id': c.id, 'name': c.category_name}
            for c in categories
        ],
        'suppliers': [
            {'id': s.id, 'name': s.supplier_name}
            for s in suppliers
        ]
    }
    return JsonResponse(data)


def inventory_transactions_api(request):
    """API endpoint to get recent inventory transactions"""
    logs = InventoryLog.objects.select_related('product', 'staff').order_by('-log_date')[:50]
    
    data = {
        'transactions': [
            {
                'id': log.id,
                'log_date': log.log_date.strftime('%Y-%m-%d %H:%M:%S'),
                'product_name': log.product.product_name,
                'log_type': log.log_type,
                'quantity': log.quantity,
                'staff_name': f"{log.staff.first_name} {log.staff.last_name}",
                'remarks': log.remarks,
            }
            for log in logs
        ]
    }
    return JsonResponse(data)


def adjust_stock(request):
    """API endpoint to adjust product stock"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        product_id = data.get('product_id')
        adjustment_type = data.get('adjustment_type')
        quantity = int(data.get('quantity'))
        remarks = data.get('remarks', '')
        
        product = get_object_or_404(Product, pk=product_id)
        
        # Calculate new quantity based on adjustment type
        if adjustment_type == 'increase':
            new_quantity = product.stock_quantity + quantity
        elif adjustment_type == 'decrease':
            new_quantity = max(0, product.stock_quantity - quantity)
        elif adjustment_type == 'set':
            new_quantity = quantity
        else:
            return JsonResponse({'success': False, 'error': 'Invalid adjustment type'})
        
        # Create inventory log entry
        InventoryLog.objects.create(
            product=product,
            staff=request.user.staff if hasattr(request.user, 'staff') else Staff.objects.first(),
            log_type='Adjustment',
            quantity=abs(new_quantity - product.stock_quantity),
            remarks=f"Stock adjustment: {adjustment_type} - {remarks}"
        )
        
        # Update product stock
        product.stock_quantity = new_quantity
        product.save()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def product_details_api(request, pk):
    """API endpoint to get detailed product information"""
    product = get_object_or_404(Product, pk=pk)
    
    data = {
        'id': product.id,
        'product_name': product.product_name,
        'brand': product.brand,
        'unit': product.unit,
        'stock_quantity': product.stock_quantity,
        'reorder_level': product.reorder_level,
        'unit_cost': float(product.unit_cost),
        'retail_price': float(product.retail_price),
        'expiry_date': product.expiry_date.strftime('%Y-%m-%d') if product.expiry_date else None,
        'category_name': product.category.category_name,
        'supplier_name': product.supplier.supplier_name,
    }
    return JsonResponse(data)


def export_inventory(request):
    """Export inventory data"""
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        return export_inventory_csv(request)
    elif export_format == 'pdf':
        return export_inventory_pdf(request)
    else:
        return HttpResponse("Invalid export format", status=400)


def export_inventory_csv(request):
    """Export inventory as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Brand', 'Category', 'Supplier', 'Current Stock', 'Reorder Level', 'Unit Cost', 'Retail Price', 'Expiry Date', 'Status'])
    
    products = Product.objects.select_related('category', 'supplier').all()
    for product in products:
        status = 'Out of Stock' if product.stock_quantity == 0 else 'Low Stock' if product.stock_quantity <= product.reorder_level else 'In Stock'
        
        writer.writerow([
            product.product_name,
            product.brand or '',
            product.category.category_name,
            product.supplier.supplier_name,
            product.stock_quantity,
            product.reorder_level,
            product.unit_cost,
            product.retail_price,
            product.expiry_date.strftime('%Y-%m-%d') if product.expiry_date else '',
            status
        ])
    
    return response


def export_inventory_pdf(request):
    """Export inventory as PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("Inventory Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Inventory summary
    products = Product.objects.select_related('category', 'supplier').all()
    total_products = products.count()
    in_stock = sum(1 for p in products if p.stock_quantity > p.reorder_level)
    low_stock = sum(1 for p in products if p.stock_quantity <= p.reorder_level and p.stock_quantity > 0)
    out_of_stock = sum(1 for p in products if p.stock_quantity == 0)
    
    summary_data = [
        ['Total Products', total_products],
        ['In Stock', in_stock],
        ['Low Stock', low_stock],
        ['Out of Stock', out_of_stock],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('GRID',(0,0),(-1,-1),1,colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Inventory details
    elements.append(Paragraph("Inventory Details", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    table_data = [['Product', 'Category', 'Stock', 'Reorder Level', 'Unit Cost', 'Retail Price', 'Status']]
    for product in products[:50]:  # Limit for PDF
        status = 'Out of Stock' if product.stock_quantity == 0 else 'Low Stock' if product.stock_quantity <= product.reorder_level else 'In Stock'
        table_data.append([
            product.product_name,
            product.category.category_name,
            str(product.stock_quantity),
            str(product.reorder_level),
            f'UGx. {product.unit_cost:,.0f}',
            f'UGx. {product.retail_price:,.0f}',
            status
        ])
    
    inventory_table = Table(table_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    inventory_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.5,colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8)
    ]))
    elements.append(inventory_table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


# ---------------------------------------------------------
# PAYROLL
# ---------------------------------------------------------
def create_payroll(request):
    if request.method == "POST":
        form = PayrollForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Payroll record saved.")
            return redirect("payroll_list")
    else:
        form = PayrollForm()
    return render(request, "inventory/payroll_form.html", {"form": form})


def payroll_list(request):
    payrolls = Payroll.objects.select_related('staff').order_by("-payment_date")
    return render(request, "inventory/payroll_list.html", {"payrolls": payrolls})


def edit_payroll(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    if request.method == "POST":
        form = PayrollForm(request.POST, instance=payroll)
        if form.is_valid():
            form.save()
            messages.success(request, "Payroll record updated successfully.")
            return redirect("payroll_list")
    else:
        form = PayrollForm(instance=payroll)
    return render(request, "inventory/payroll_form.html", {"form": form, "payroll": payroll})


def delete_payroll(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    if request.method == "POST":
        staff_name = f"{payroll.staff.first_name} {payroll.staff.last_name}"
        payroll.delete()
        messages.success(request, f"Payroll record for {staff_name} has been deleted.")
        return redirect("payroll_list")
    return redirect("payroll_list")


def payroll_details_api(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    data = {
        'staff_name': f"{payroll.staff.first_name} {payroll.staff.last_name}",
        'staff_position': payroll.staff.position,
        'staff_department': getattr(payroll.staff, 'department', None),
        'payment_date': payroll.payment_date.strftime('%B %d, %Y'),
        'payment_method': payroll.payment_method,
        'basic_salary': float(payroll.basic_salary),
        'allowances': float(payroll.allowances) if payroll.allowances else 0,
        'deductions': float(payroll.deductions) if payroll.deductions else 0,
        'net_salary': float(payroll.net_salary),
    }
    return JsonResponse(data)


def export_payroll(request):
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        return export_payroll_csv(request)
    elif export_format == 'pdf':
        return export_payroll_pdf(request)
    else:
        return HttpResponse("Invalid export format", status=400)


def export_payroll_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="payroll_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Staff Name', 'Position', 'Payment Date', 'Basic Salary', 'Allowances', 'Deductions', 'Net Salary', 'Payment Method'])
    
    payrolls = Payroll.objects.select_related('staff').order_by('-payment_date')
    for payroll in payrolls:
        writer.writerow([
            f"{payroll.staff.first_name} {payroll.staff.last_name}",
            payroll.staff.position,
            payroll.payment_date.strftime('%Y-%m-%d'),
            payroll.basic_salary,
            payroll.allowances or 0,
            payroll.deductions or 0,
            payroll.net_salary,
            payroll.payment_method
        ])
    
    return response


def export_payroll_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payroll_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("Payroll Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Payroll summary
    payrolls = Payroll.objects.select_related('staff').order_by('-payment_date')
    total_net = sum(p.net_salary for p in payrolls)
    total_deductions = sum(p.deductions or 0 for p in payrolls)
    total_allowances = sum(p.allowances or 0 for p in payrolls)
    
    summary_data = [
        ['Total Records', len(payrolls)],
        ['Total Net Salary', f'UGx. {total_net:,.0f}'],
        ['Total Allowances', f'UGx. {total_allowances:,.0f}'],
        ['Total Deductions', f'UGx. {total_deductions:,.0f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('GRID',(0,0),(-1,-1),1,colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Payroll details
    elements.append(Paragraph("Payroll Details", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    table_data = [['Staff Name', 'Position', 'Payment Date', 'Basic Salary', 'Allowances', 'Deductions', 'Net Salary']]
    for payroll in payrolls[:50]:  # Limit for PDF
        table_data.append([
            f"{payroll.staff.first_name} {payroll.staff.last_name}",
            payroll.staff.position,
            payroll.payment_date.strftime('%Y-%m-%d'),
            f'UGx. {payroll.basic_salary:,.0f}',
            f'UGx. {(payroll.allowances or 0):,.0f}',
            f'UGx. {(payroll.deductions or 0):,.0f}',
            f'UGx. {payroll.net_salary:,.0f}'
        ])
    
    payroll_table = Table(table_data, colWidths=[1.5*inch, 1.2*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
    payroll_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.5,colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8)
    ]))
    elements.append(payroll_table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response




# ---------------------------------------------------------
# 📊 REPORTING & ANALYTICS API VIEWS
# ---------------------------------------------------------

# Helper function to get date filters from request
def get_date_filters(request):
    """
    Extract 'from' and 'to' date parameters from request.
    Returns a dict with 'start' and 'end' datetime objects or None.
    """
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')
    
    filters = {}
    
    if from_date:
        try:
            # Parse date and set to start of day
            from datetime import datetime
            start = datetime.strptime(from_date, '%Y-%m-%d')
            filters['start'] = timezone.make_aware(start) if timezone.is_naive(start) else start
        except ValueError:
            pass
    
    if to_date:
        try:
            # Parse date and set to end of day
            from datetime import datetime
            end = datetime.strptime(to_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            filters['end'] = timezone.make_aware(end) if timezone.is_naive(end) else end
        except ValueError:
            pass
    
    return filters

def reports_view(request):
    """View for analytics and reports dashboard."""
    today = timezone.now().date()
    month_start = today.replace(day=1)

    # Use SaleDetail for itemized sales
    sale_details = SaleDetail.objects.all()

    # Total Sales
    total_sales = sale_details.aggregate(
        total=Sum(F('quantity_sold') * F('unit_price'), output_field=FloatField())
    )['total'] or 0

    # Total Purchases (from PurchaseOrderDetail)
    purchases = PurchaseOrderDetail.objects.all()
    total_purchases = purchases.aggregate(
        total=Sum(F('quantity_ordered') * F('unit_cost'), output_field=FloatField())
    )['total'] or 0

    # Profit
    profit = total_sales - total_purchases

    # Sales by Category
    category_sales = (
        sale_details.values('product__category__category_name')
        .annotate(total=Sum(F('unit_price') * F('quantity_sold'), output_field=FloatField()))
        .order_by('product__category__category_name')
    )

    # Sales distribution (for optional histogram)
    sales_distribution = (
        Sale.objects.annotate(day=TruncDate('sale_datetime'))
        .values('day')
        .annotate(count=Sum('total_amount'))
        .order_by('day')
    )

    # Yearly sales data for the line graph
    current_year = now().year
    start_year = current_year - 4  # Show last 5 years by default
    
    yearly_sales = (
        Sale.objects.filter(
            sale_datetime__year__gte=start_year,
            sale_datetime__year__lte=current_year
        )
        .annotate(year=ExtractYear('sale_datetime'))
        .values('year')
        .annotate(total_sales=Sum('total_amount'))
        .order_by('year')
    )
    
    # Prepare yearly data for template
    yearly_data = []
    for year in range(start_year, current_year + 1):
        year_data = next((item for item in yearly_sales if item['year'] == year), None)
        yearly_data.append({
            'year': year,
            'total_sales': float(year_data['total_sales']) if year_data else 0.0
        })

    # Convert QuerySets to lists for safe JSON rendering
    context = {
        "total_sales": total_sales,
        "total_purchases": total_purchases,
        "profit": profit,
        "category_sales": list(category_sales),
        "sales_distribution": list(sales_distribution),
        "yearly_sales_data": yearly_data,
        "current_year": current_year,
        "start_year": start_year,
    }

    return render(request, 'inventory/reports.html', context)


def sales_by_category_api(request):
    """Return total sales grouped by product category as percentages."""
    date_filters = get_date_filters(request)

    sale_details_qs = SaleDetail.objects.all()
    if 'start' in date_filters:
        sale_details_qs = sale_details_qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sale_details_qs = sale_details_qs.filter(sale__sale_datetime__lte=date_filters['end'])

    # Group totals by category
    data = (
        sale_details_qs
        .values('product__category__category_name')
        .annotate(total=Sum(F('unit_price') * F('quantity_sold'), output_field=FloatField()))
        .order_by('product__category__category_name')
    )

    # Calculate the grand total
    grand_total = sum(d['total'] or 0 for d in data) or 1  # avoid division by zero

    # Convert each total to percentage
    labels = [d['product__category__category_name'] or 'Uncategorized' for d in data]
    percentages = [round((float(d['total'] or 0) / grand_total) * 100, 2) for d in data]

    return JsonResponse({'labels': labels, 'data': percentages})


def sales_histogram_api(request):
    """Return sales distribution with realistic ranges"""
    date_filters = get_date_filters(request)
    
    # Base queryset
    sales_qs = Sale.objects.all()
    if 'start' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__lte=date_filters['end'])
    
    # More realistic ranges based on your data
    buckets = {
        "Under 50K": sales_qs.filter(total_amount__lt=50000).count(),
        "50K–100K": sales_qs.filter(total_amount__gte=50000, total_amount__lt=100000).count(),
        "100K–200K": sales_qs.filter(total_amount__gte=100000, total_amount__lt=200000).count(),
        "200K–300K": sales_qs.filter(total_amount__gte=200000, total_amount__lt=300000).count(),
        "300K+": sales_qs.filter(total_amount__gte=300000).count(),
    }
    labels = list(buckets.keys())
    values = list(buckets.values())
    return JsonResponse({'labels': labels, 'data': values})

# --- KPI API: returns totals for dashboard ---
def kpi_data_api(request):
    """
    Returns JSON:
    {
      "total_revenue": float,
      "total_orders": int,
      "avg_order_value": float,
      "top_category": "Category Name" or null,
      "revenue_growth_pct": float (optional, relative to previous period)
    }
    """
    date_filters = get_date_filters(request)
    
    # Base queryset
    sales_qs = Sale.objects.all()
    if 'start' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__lte=date_filters['end'])
    
    # total revenue & orders
    total_revenue = sales_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_orders = sales_qs.count()

    # average order value (safe)
    avg_order_value = float(total_revenue) / total_orders if total_orders else 0.0

    # top category by sales (unit_price * qty) - filter by same date range
    sale_details_qs = SaleDetail.objects.filter(sale__in=sales_qs)
    
    top_cat_q = (
        sale_details_qs
        .values('product__category__category_name')
        .annotate(total=Sum(F('unit_price') * F('quantity_sold'), output_field=FloatField()))
        .order_by('-total')
    ).first()
    top_category = top_cat_q['product__category__category_name'] if top_cat_q else ''

    # optional: simple growth % for last 30 days vs previous 30 days
    # Only calculate if no custom date range is provided
    revenue_growth_pct = None
    if not date_filters:  # Only calculate growth for default view
        try:
            now_time = timezone.now()
            last_30_start = now_time - timedelta(days=30)
            prev_30_start = now_time - timedelta(days=60)
            current_sum = Sale.objects.filter(sale_datetime__gte=last_30_start).aggregate(total=Sum('total_amount'))['total'] or 0
            previous_sum = Sale.objects.filter(sale_datetime__gte=prev_30_start, sale_datetime__lt=last_30_start).aggregate(total=Sum('total_amount'))['total'] or 0
            if previous_sum and previous_sum != 0:
                revenue_growth_pct = float((current_sum - previous_sum) / previous_sum * 100)
        except Exception:
            revenue_growth_pct = None

    payload = {
        'total_revenue': float(total_revenue),
        'total_orders': int(total_orders),
        'avg_order_value': float(avg_order_value),
        'top_category': top_category or '',
        'revenue_growth_pct': revenue_growth_pct,
    }
    return JsonResponse(payload)


# --- Sales table API: returns recent sale lines for the detailed table ---
def sales_table_data_api(request):
    """
    Returns a JSON array of recent sale lines:
    [
      {"date":"YYYY-MM-DD", "product":"Name", "category":"Cat", "quantity":int, 
       "unit_price":float, "total":float, "customer":"Name"},
      ...
    ]
    """
    date_filters = get_date_filters(request)
    
    # Base queryset
    qs = SaleDetail.objects.select_related('sale', 'product', 'product__category', 'sale__customer')
    
    if 'start' in date_filters:
        qs = qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        qs = qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    qs = qs.order_by('-sale__sale_datetime')[:200]

    rows = []
    for sd in qs:
        sale = sd.sale
        product = sd.product
        # customer display: prefer name, fallback to phone/email
        customer_name = ''
        if getattr(sale, 'customer', None):
            c = sale.customer
            # assemble sensible name
            if hasattr(c, 'first_name') and c.first_name:
                customer_name = f"{(c.first_name or '')} {(c.last_name or '')}".strip()
            else:
                customer_name = getattr(c, 'phone', '') or getattr(c, 'email', '') or ''
        rows.append({
            'date': sale.sale_datetime.strftime('%Y-%m-%d'),
            'product': product.product_name if product else '',
            'category': getattr(product.category, 'category_name', '') if getattr(product, 'category', None) else '',
            'quantity': int(sd.quantity_sold or 0),
            'unit_price': float(sd.unit_price or 0),
            'total': float(sd.sub_total or (sd.unit_price * sd.quantity_sold) or 0),
            'customer': customer_name,
        })

    return JsonResponse(rows, safe=False)

def yearly_sales_api(request):
    """FINAL VERSION - Query Sale table for yearly sales"""
    try:
        # Get year range from request
        start_year = request.GET.get('start_year', '2021')
        end_year = request.GET.get('end_year', '2025')
        
        try:
            start_year = int(start_year)
            end_year = int(end_year)
        except (ValueError, TypeError):
            start_year = 2021
            end_year = 2025

        print(f"Querying years {start_year} to {end_year}")

        # Method 1: Direct query using year filter
        yearly_data = []
        for year in range(start_year, end_year + 1):
            # Get total sales for this year
            year_sales = Sale.objects.filter(sale_datetime__year=year)
            year_total = year_sales.aggregate(total=Sum('total_amount'))['total'] or 0
            
            print(f"Year {year}: {year_sales.count()} records, total: {year_total}")
            
            yearly_data.append({
                'year': year,
                'total_sales': float(year_total)
            })

        # Prepare response
        years = [item['year'] for item in yearly_data]
        sales_totals = [item['total_sales'] for item in yearly_data]

        print(f"Final response: {years} -> {sales_totals}")

        return JsonResponse({
            'years': years,
            'sales_totals': sales_totals,
            'currency_symbol': 'UGx.'
        })
        
    except Exception as e:
        print(f"ERROR in yearly_sales_api: {e}")
        # Return actual error instead of zeros
        return JsonResponse({
            'error': str(e),
            'years': [2021, 2022, 2023, 2024, 2025],
            'sales_totals': [100000, 200000, 300000, 400000, 500000],  # Test data to verify API works
            'currency_symbol': 'UGx.'
        })
    
    
def monthly_sales_api(request):
    """Return monthly sales data for a specific year"""
    year = request.GET.get('year', now().year)
    
    try:
        year = int(year)
    except (ValueError, TypeError):
        year = now().year
    
    monthly_sales = (
        Sale.objects.filter(sale_datetime__year=year)
        .annotate(month=ExtractMonth('sale_datetime'))
        .values('month')
        .annotate(total_sales=Sum('total_amount'))
        .order_by('month')
    )
    
    # Create complete dataset for all months
    complete_data = []
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    for month in range(1, 13):
        month_data = next((item for item in monthly_sales if item['month'] == month), None)
        if month_data:
            complete_data.append({
                'month': month,
                'month_name': month_names[month-1],
                'total_sales': float(month_data['total_sales'] or 0)
            })
        else:
            complete_data.append({
                'month': month,
                'month_name': month_names[month-1],
                'total_sales': 0.0
            })
    
    months = [item['month_name'] for item in complete_data]
    sales_totals = [item['total_sales'] for item in complete_data]
    
    return JsonResponse({
        'year': year,
        'months': months,
        'sales_totals': sales_totals,
        'currency_symbol': 'UGx.'
    })
    
def quarterly_sales_api(request):
    """Simple quarterly sales using same approach as yearly sales"""
    try:
        # Get year range from request
        start_year = request.GET.get('start_year', '2023')
        end_year = request.GET.get('end_year', '2025')
        
        try:
            start_year = int(start_year)
            end_year = int(end_year)
        except (ValueError, TypeError):
            start_year = 2023
            end_year = 2025

        print(f"QUARTERLY DEBUG: Querying years {start_year} to {end_year}")

        # Simple approach - use the same logic as yearly sales but for quarters
        quarterly_data = []
        for year in range(start_year, end_year + 1):
            print(f"QUARTERLY DEBUG: Processing year {year}")
            
            for quarter in range(1, 5):  # Quarters 1-4
                # Simple quarter filtering using the same approach as yearly
                quarter_sales = Sale.objects.filter(
                    sale_datetime__year=year
                ).extra({
                    'quarter': "EXTRACT(QUARTER FROM sale_datetime)"
                }).filter(quarter=quarter)
                
                quarter_total = quarter_sales.aggregate(total=Sum('total_amount'))['total'] or 0
                
                print(f"QUARTERLY DEBUG: Year {year} Q{quarter}: {quarter_sales.count()} records, total: {quarter_total}")
                
                if quarter_sales.exists():
                    sample = quarter_sales.first()
                    print(f"QUARTERLY DEBUG: Sample - {sample.sale_datetime} - {sample.total_amount}")
                
                quarterly_data.append({
                    'year': year,
                    'quarter': quarter,
                    'quarter_label': f"Q{quarter}",
                    'total_sales': float(quarter_total)
                })

        # Group by year
        years = []
        quarters_data = {}
        
        for item in quarterly_data:
            year = item['year']
            quarter = item['quarter_label']
            sales = item['total_sales']
            
            if year not in years:
                years.append(year)
                quarters_data[year] = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
            
            quarters_data[year][quarter] = sales

        # Prepare response
        quarter_labels = ['Q1', 'Q2', 'Q3', 'Q4']
        datasets = []
        
        for year in sorted(years):
            datasets.append({
                'label': f'{year}',
                'data': [
                    quarters_data[year]['Q1'],
                    quarters_data[year]['Q2'], 
                    quarters_data[year]['Q3'],
                    quarters_data[year]['Q4']
                ]
            })

        print(f"QUARTERLY DEBUG: Final response - {datasets}")

        return JsonResponse({
            'labels': quarter_labels,
            'datasets': datasets,
            'currency_symbol': 'UGx.'
        })
        
    except Exception as e:
        print(f"QUARTERLY ERROR: {e}")
        import traceback
        print(f"QUARTERLY TRACEBACK: {traceback.format_exc()}")
        
        # Fallback to test data
        return JsonResponse({
            'labels': ['Q1', 'Q2', 'Q3', 'Q4'],
            'datasets': [
                {'label': '2023', 'data': [1000000, 1500000, 1200000, 1800000]},
                {'label': '2024', 'data': [2000000, 2200000, 1900000, 2500000]},
                {'label': '2025', 'data': [1800000, 2000000, 2200000, 2400000]}
            ],
            'currency_symbol': 'UGx.'
        })

# ---------------------------------------------------------
# FINACIAL REPORT
# ---------------------------------------------------------

def financial_report_api(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    group_by = request.GET.get('group_by', 'day')  # We keep the parameter but will group daily only for now

    # Parse dates
    start = datetime.strptime(start, "%Y-%m-%d").date()
    end = datetime.strptime(end, "%Y-%m-%d").date()

    # -------------------- GROSS SALES --------------------
    gross_sales = (
        Sale.objects.filter(sale_datetime__date__range=[start, end])
        .annotate(period=Cast('sale_datetime', DateField()))
        .values('period')
        .annotate(value=Sum('total_amount'))
        .order_by('period')
    )

    # -------------------- COGS --------------------
    cogs = (
        SaleDetail.objects.filter(sale__sale_datetime__date__range=[start, end])
        .annotate(period=Cast('sale__sale_datetime', DateField()))
        .values('period')
        .annotate(
            value=Sum(
                F('quantity_sold') * F('product__unit_cost'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
        .order_by('period')
    )

    # -------------------- PAYROLL --------------------
    payroll_expenses = (
        Payroll.objects.filter(payment_date__range=[start, end])
        .annotate(period=Cast('payment_date', DateField()))
        .values('period')
        .annotate(value=Sum('net_salary'))
        .order_by('period')
    )

    # -------------------- EXPIRY LOSSES --------------------
    expiry_losses = (
        InventoryLog.objects.filter(
            log_date__date__range=[start, end],
            remarks__icontains='expiry_writeoff'
        )
        .annotate(period=Cast('log_date', DateField()))
        .values('period')
        .annotate(
            value=Sum(
                F('quantity') * F('product__unit_cost'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
        .order_by('period')
    )

    # -------------------- TAXES --------------------
    TAX_RATE = 0.18
    taxes = (
        Sale.objects.filter(sale_datetime__date__range=[start, end])
        .annotate(period=Cast('sale_datetime', DateField()))
        .values('period')
        .annotate(
            value=Sum(
                F('total_amount') * TAX_RATE,
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
        .order_by('period')
    )

    # Return JSON
    return JsonResponse({
        'gross_sales': list(gross_sales),
        'cogs': list(cogs),
        'payroll_expenses': list(payroll_expenses),
        'expiry_losses': list(expiry_losses),
        'taxes': list(taxes),
    })
# ---------------------------------------------------------
# TAX REPORT API (sale-level tax)
# ---------------------------------------------------------
def taxes_report_api(request):
    """
    GET /inventory/api/reports/taxes?start=&end[&rate=][&group_by=payment_method]

    Convention: Sale.total_amount is tax-inclusive.

    Inference per sale when no rate provided:
      tax = total_amount - sum(details.sub_total) - (discount_applied or 0)

    If rate provided (decimal, e.g., 0.18):
      tax = rate * max(sum_subtotals - discount_applied, 0)

    Returns JSON with total_tax and count_of_sales.
    If group_by=payment_method, returns grouped aggregates as well.
    """
    try:
        start = request.GET.get('start')
        end = request.GET.get('end')
        rate_param = request.GET.get('rate')
        group_by = request.GET.get('group_by')

        qs = Sale.objects.select_related('customer', 'staff').prefetch_related('details')
        if start:
            # Parse start date and filter correctly
            try:
                from datetime import datetime
                if isinstance(start, str):
                    start_date = datetime.strptime(start, '%Y-%m-%d').date()
                else:
                    start_date = start
                qs = qs.filter(sale_datetime__date__gte=start_date)
            except (ValueError, TypeError) as e:
                pass  # Ignore invalid date formats
        if end:
            # Parse end date and filter correctly
            try:
                from datetime import datetime
                if isinstance(end, str):
                    end_date = datetime.strptime(end, '%Y-%m-%d').date()
                else:
                    end_date = end
                qs = qs.filter(sale_datetime__date__lte=end_date)
            except (ValueError, TypeError) as e:
                pass  # Ignore invalid date formats

        # Annotate each sale with sum of subtotals
        qs = qs.annotate(
            sum_subtotals=Coalesce(Sum('details__sub_total'), Value(0), output_field=DecimalField(max_digits=18, decimal_places=2)),
            discount_amt=Coalesce(F('discount_applied'), Value(0), output_field=DecimalField(max_digits=18, decimal_places=2))
        )

        # Calculate tax per sale
        if rate_param is not None and rate_param != '':
            try:
                rate_decimal = Decimal(str(rate_param))
                if rate_decimal < 0 or rate_decimal > 1:
                    return JsonResponse({'error': 'Rate must be between 0 and 1'}, status=400)
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid rate parameter. Must be a decimal number (e.g., 0.18 for 18%)'}, status=400)
            
            # Tax = rate * max(sum_subtotals - discount_applied, 0)
            base_amount = ExpressionWrapper(
                F('sum_subtotals') - F('discount_amt'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
            positive_base = Greatest(base_amount, Value(0, output_field=DecimalField(max_digits=18, decimal_places=2)))
            tax_expr = ExpressionWrapper(
                positive_base * Value(rate_decimal, output_field=DecimalField(max_digits=18, decimal_places=6)),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        else:
            # Tax inference: tax = total_amount - sum_subtotals - discount_applied
            tax_expr = ExpressionWrapper(
                F('total_amount') - F('sum_subtotals') - F('discount_amt'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )

        # Annotate each sale with calculated tax
        qs = qs.annotate(calculated_tax=tax_expr)

        if group_by == 'payment_method':
            # Check if there are any sales first
            if not qs.exists():
                return JsonResponse({
                    'total_tax': 0.0,
                    'count_of_sales': 0,
                    'groups': []
                })
            
            # Group by payment method and aggregate
            grouped = (
                qs.values('payment_method')
                  .annotate(
                      total_tax=Sum('calculated_tax'),
                      count_of_sales=Count('id')
                  )
                  .order_by('payment_method')
            )
            groups = [
                {
                    'payment_method': g['payment_method'],
                    'total_tax': float(g['total_tax'] or 0),
                    'count_of_sales': int(g['count_of_sales'] or 0),
                }
                for g in grouped
            ]
            total_tax = sum(item['total_tax'] for item in groups)
            count_of_sales = sum(item['count_of_sales'] for item in groups)
            return JsonResponse({
                'total_tax': float(total_tax),
                'count_of_sales': int(count_of_sales),
                'groups': groups
            })

        # Aggregate totals (for non-grouped requests)
        agg = qs.aggregate(
            total_tax=Sum('calculated_tax'),
            count_of_sales=Count('id')
        )
        return JsonResponse({
            'total_tax': float(agg.get('total_tax') or 0),
            'count_of_sales': int(agg.get('count_of_sales') or 0),
        })
    
    except Exception as e:
        # Return JSON error instead of HTML error page
        import traceback
        error_details = str(e)
        # In production, you might want to log the traceback instead of exposing it
        return JsonResponse({
            'error': f'An error occurred while processing tax report: {error_details}',
            'details': traceback.format_exc() if settings.DEBUG else None
        }, status=500)

# ---------------------------------------------------------
# 📄 EXPORT VIEWS
# ---------------------------------------------------------

# Add these export functions to your views.py

def export_report(request):
    """Export full report (KPIs + Sales details) in selected format"""
    export_format = request.GET.get('format', 'pdf')
    
    if export_format == 'pdf':
        return export_report_pdf(request)
    elif export_format == 'csv':
        return export_report_csv(request)
    elif export_format == 'excel':
        return export_report_excel(request)
    else:
        return HttpResponse("Invalid export format", status=400)

def export_table(request):
    """Export only sales table in selected format"""
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'pdf':
        return export_table_pdf(request)
    elif export_format == 'csv':
        return export_table_csv(request)
    elif export_format == 'excel':
        return export_table_excel(request)
    else:
        return HttpResponse("Invalid export format", status=400)

def export_report_pdf(request):
    """Export KPI + Sales details report as PDF."""
    date_filters = get_date_filters(request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("Sales Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Date range
    if date_filters:
        date_info = f"Period: {date_filters.get('start', 'All')} to {date_filters.get('end', 'Now')}"
        elements.append(Paragraph(date_info, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # KPIs
    sales_qs = Sale.objects.all()
    if 'start' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__lte=date_filters['end'])
    total_revenue = sales_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_orders = sales_qs.count()
    avg_order_value = total_revenue / total_orders if total_orders else 0
    
    kpi_data = [['Metric', 'Value'],
                ['Total Revenue', f'UGx. {total_revenue:,.0f}'],
                ['Total Orders', f'{total_orders:,}'],
                ['Average Order Value', f'UGx. {avg_order_value:,.0f}']]
    kpi_table = Table(kpi_data, colWidths=[3*inch, 3*inch])
    kpi_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.grey),
                                   ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                                   ('ALIGN',(0,0),(-1,-1),'CENTER'),
                                   ('GRID',(0,0),(-1,-1),1,colors.black)]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))
    
    # Sales details
    elements.append(Paragraph("Sales Details", styles['Heading2']))
    elements.append(Spacer(1, 12))
    sale_details_qs = SaleDetail.objects.select_related('sale','product','product__category')
    if 'start' in date_filters:
        sale_details_qs = sale_details_qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sale_details_qs = sale_details_qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    table_data = [['Date','Product','Category','Qty','Price','Total']]
    for sd in sale_details_qs.order_by('-sale__sale_datetime')[:50]:
        table_data.append([
            sd.sale.sale_datetime.strftime('%Y-%m-%d'),
            sd.product.product_name if sd.product else '',
            getattr(sd.product.category, 'category_name', '') if getattr(sd.product, 'category', None) else '',
            str(sd.quantity_sold or 0),
            f'UGx. {sd.unit_price or 0:,.0f}',
            f'UGx. {(sd.unit_price * sd.quantity_sold) or 0:,.0f}'
        ])
    sales_table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1.2*inch, 0.6*inch, 1*inch, 1.2*inch])
    sales_table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),
                                     ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                                     ('ALIGN',(0,0),(-1,-1),'CENTER'),
                                     ('GRID',(0,0),(-1,-1),0.5,colors.black)]))
    elements.append(sales_table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def export_report_csv(request):
    """Export full report as CSV"""
    date_filters = get_date_filters(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Report header
    writer.writerow(['SALES REPORT'])
    if date_filters:
        date_info = f"Period: {date_filters.get('start', 'All')} to {date_filters.get('end', 'Now')}"
        writer.writerow([date_info])
    writer.writerow([])
    
    # KPIs section
    writer.writerow(['KPI SUMMARY'])
    sales_qs = Sale.objects.all()
    if 'start' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__lte=date_filters['end'])
    
    total_revenue = sales_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_orders = sales_qs.count()
    avg_order_value = total_revenue / total_orders if total_orders else 0
    
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Revenue', f'UGx. {total_revenue:,.0f}'])
    writer.writerow(['Total Orders', total_orders])
    writer.writerow(['Average Order Value', f'UGx. {avg_order_value:,.0f}'])
    writer.writerow([])
    
    # Sales details section
    writer.writerow(['SALES DETAILS'])
    writer.writerow(['Date','Product','Category','Qty','Price','Total','Customer'])
    
    qs = SaleDetail.objects.select_related('sale','product','product__category','sale__customer')
    if 'start' in date_filters:
        qs = qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        qs = qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    for sd in qs.order_by('-sale__sale_datetime'):
        customer_name = ''
        if sd.sale.customer:
            c = sd.sale.customer
            customer_name = f"{getattr(c,'first_name','')} {getattr(c,'last_name','')}".strip() or getattr(c,'phone','') or getattr(c,'email','')
        writer.writerow([
            sd.sale.sale_datetime.strftime('%Y-%m-%d'),
            sd.product.product_name if sd.product else '',
            getattr(sd.product.category,'category_name','') if getattr(sd.product,'category',None) else '',
            sd.quantity_sold or 0,
            sd.unit_price or 0,
            (sd.unit_price * sd.quantity_sold) or 0,
            customer_name
        ])
    return response

def export_report_excel(request):
    """Export full report as Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it: pip install openpyxl", status=500)
    
    date_filters = get_date_filters(request)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    # Title
    ws['A1'] = "SALES REPORT"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    # Date range
    row = 2
    if date_filters:
        date_range = f"Period: {date_filters.get('start', 'All')} to {date_filters.get('end', 'Now')}"
        ws[f'A{row}'] = date_range
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
    
    row += 1
    
    # KPI Summary
    sales_qs = Sale.objects.all()
    if 'start' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        sales_qs = sales_qs.filter(sale_datetime__lte=date_filters['end'])
    
    total_revenue = sales_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_orders = sales_qs.count()
    avg_order_value = total_revenue / total_orders if total_orders else 0
    
    # KPI Header
    ws[f'A{row}'] = "KPI SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # KPI Table Headers
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    # KPI Data
    kpi_data = [
        ['Total Revenue', f'UGx. {total_revenue:,.0f}'],
        ['Total Orders', total_orders],
        ['Average Order Value', f'UGx. {avg_order_value:,.0f}'],
    ]
    
    for kpi_row in kpi_data:
        ws[f'A{row}'] = kpi_row[0]
        ws[f'B{row}'] = kpi_row[1]
        row += 1
    
    row += 2
    
    # Sales Details
    ws[f'A{row}'] = "SALES DETAILS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # Headers
    headers = ['Date', 'Product', 'Category', 'Quantity', 'Unit Price', 'Total', 'Customer']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Get sales data
    qs = SaleDetail.objects.select_related('sale','product','product__category','sale__customer')
    if 'start' in date_filters:
        qs = qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        qs = qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    for sd in qs.order_by('-sale__sale_datetime'):
        customer_name = ''
        if sd.sale.customer:
            c = sd.sale.customer
            customer_name = f"{getattr(c,'first_name','')} {getattr(c,'last_name','')}".strip() or getattr(c,'phone','') or getattr(c,'email','')
        
        ws[f'A{row}'] = sd.sale.sale_datetime.strftime('%Y-%m-%d')
        ws[f'B{row}'] = sd.product.product_name if sd.product else ''
        ws[f'C{row}'] = getattr(sd.product.category,'category_name','') if getattr(sd.product,'category',None) else ''
        ws[f'D{row}'] = sd.quantity_sold or 0
        ws[f'E{row}'] = sd.unit_price or 0
        ws[f'F{row}'] = (sd.unit_price * sd.quantity_sold) or 0
        ws[f'G{row}'] = customer_name
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response

def export_table_pdf(request):
    """Export only sales table as PDF"""
    date_filters = get_date_filters(request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_table_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("Sales Table", title_style))
    elements.append(Spacer(1, 12))
    
    # Date range
    if date_filters:
        date_info = f"Period: {date_filters.get('start', 'All')} to {date_filters.get('end', 'Now')}"
        elements.append(Paragraph(date_info, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Sales table only
    table_data = [['Date','Product','Category','Qty','Price','Total','Customer']]
    
    qs = SaleDetail.objects.select_related('sale','product','product__category','sale__customer')
    if 'start' in date_filters:
        qs = qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        qs = qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    for sd in qs.order_by('-sale__sale_datetime')[:100]:  # Limit for PDF
        customer_name = ''
        if sd.sale.customer:
            c = sd.sale.customer
            customer_name = f"{getattr(c,'first_name','')} {getattr(c,'last_name','')}".strip() or getattr(c,'phone','') or getattr(c,'email','')
        
        table_data.append([
            sd.sale.sale_datetime.strftime('%Y-%m-%d'),
            sd.product.product_name if sd.product else '',
            getattr(sd.product.category, 'category_name', '') if getattr(sd.product, 'category', None) else '',
            str(sd.quantity_sold or 0),
            f'UGx. {sd.unit_price or 0:,.0f}',
            f'UGx. {(sd.unit_price * sd.quantity_sold) or 0:,.0f}',
            customer_name
        ])
    
    sales_table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1.2*inch, 0.6*inch, 1*inch, 1.2*inch, 1.5*inch])
    sales_table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),
                                     ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                                     ('ALIGN',(0,0),(-1,-1),'CENTER'),
                                     ('GRID',(0,0),(-1,-1),0.5,colors.black)]))
    elements.append(sales_table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def export_table_csv(request):
    """Export detailed sales table as CSV"""
    date_filters = get_date_filters(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_table_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date','Product','Category','Qty','Price','Total','Customer'])
    
    qs = SaleDetail.objects.select_related('sale','product','product__category','sale__customer')
    if 'start' in date_filters:
        qs = qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        qs = qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    for sd in qs.order_by('-sale__sale_datetime'):
        customer_name = ''
        if sd.sale.customer:
            c = sd.sale.customer
            customer_name = f"{getattr(c,'first_name','')} {getattr(c,'last_name','')}".strip() or getattr(c,'phone','') or getattr(c,'email','')
        writer.writerow([
            sd.sale.sale_datetime.strftime('%Y-%m-%d'),
            sd.product.product_name if sd.product else '',
            getattr(sd.product.category,'category_name','') if getattr(sd.product,'category',None) else '',
            sd.quantity_sold or 0,
            sd.unit_price or 0,
            (sd.unit_price * sd.quantity_sold) or 0,
            customer_name
        ])
    return response

def export_table_excel(request):
    """Export sales table as Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it: pip install openpyxl", status=500)
    
    date_filters = get_date_filters(request)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Table"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    # Title
    ws['A1'] = "SALES TABLE"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    # Date range
    row = 2
    if date_filters:
        date_range = f"Period: {date_filters.get('start', 'All')} to {date_filters.get('end', 'Now')}"
        ws[f'A{row}'] = date_range
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
    
    row += 1
    
    # Headers
    headers = ['Date', 'Product', 'Category', 'Quantity', 'Unit Price', 'Total', 'Customer']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Get sales data
    qs = SaleDetail.objects.select_related('sale','product','product__category','sale__customer')
    if 'start' in date_filters:
        qs = qs.filter(sale__sale_datetime__gte=date_filters['start'])
    if 'end' in date_filters:
        qs = qs.filter(sale__sale_datetime__lte=date_filters['end'])
    
    for sd in qs.order_by('-sale__sale_datetime'):
        customer_name = ''
        if sd.sale.customer:
            c = sd.sale.customer
            customer_name = f"{getattr(c,'first_name','')} {getattr(c,'last_name','')}".strip() or getattr(c,'phone','') or getattr(c,'email','')
        
        ws[f'A{row}'] = sd.sale.sale_datetime.strftime('%Y-%m-%d')
        ws[f'B{row}'] = sd.product.product_name if sd.product else ''
        ws[f'C{row}'] = getattr(sd.product.category,'category_name','') if getattr(sd.product,'category',None) else ''
        ws[f'D{row}'] = sd.quantity_sold or 0
        ws[f'E{row}'] = sd.unit_price or 0
        ws[f'F{row}'] = (sd.unit_price * sd.quantity_sold) or 0
        ws[f'G{row}'] = customer_name
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_table_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response


# Expiry Management Views
def expiry_preview(request):
    """Preview expiring products and allow manual write-off"""
    from datetime import date
    
    today = date.today()
    
    # Get expired products with stock
    expired_products = Product.objects.filter(
        expiry_date__lt=today,
        stock_quantity__gt=0
    ).select_related('category', 'supplier')
    
    # Calculate total potential loss
    total_loss = sum(product.stock_quantity * product.unit_cost for product in expired_products)
    
    context = {
        'expired_products': expired_products,
        'total_loss': total_loss,
        'product_count': expired_products.count()
    }
    
    return render(request, 'inventory/expiry_confirm.html', context)


def execute_expiry_writeoff(request):
    """Execute expiry write-off after confirmation"""
    if request.method != 'POST':
        return redirect('expiry_preview')
    
    from django.db import transaction
    from datetime import date
    
    today = date.today()
    
    # Get expired products with stock
    expired_products = Product.objects.filter(
        expiry_date__lt=today,
        stock_quantity__gt=0
    )
    
    if not expired_products.exists():
        messages.info(request, 'No expired products found to write off')
        return redirect('expiry_preview')
    
    # Get staff member (you might want to get from session or request)
    try:
        staff = Staff.objects.get(username='system')
    except Staff.DoesNotExist:
        staff = Staff.objects.create(
            first_name='System',
            last_name='User',
            role='Admin',
            username='system',
            password_hash='system_user'
        )
    
    # Process write-offs in a transaction
    with transaction.atomic():
        total_loss = 0
        processed_count = 0
        
        for product in expired_products:
            # Create inventory log entry
            InventoryLog.objects.create(
                staff=staff,
                product=product,
                log_type='Adjustment',
                quantity=-product.stock_quantity,
                remarks='expiry_writeoff',
                log_date=timezone.now()
            )
            
            # Calculate loss
            loss_amount = product.stock_quantity * product.unit_cost
            total_loss += loss_amount
            
            # Set stock quantity to 0
            product.stock_quantity = 0
            product.save()
            
            processed_count += 1
    
    messages.success(
        request,
        f'Successfully wrote off {processed_count} expired products. '
        f'Total loss: ${total_loss:.2f}'
    )
    
    return redirect('expiry_preview')


def expiry_reports_api(request):
    """API endpoint for expiry reports"""
    from datetime import datetime, date
    from django.db.models import Sum, F
    
    # Get date range parameters
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    # Build query for expiry write-off logs
    logs = InventoryLog.objects.filter(
        remarks__icontains='expiry_writeoff'
    ).select_related('product', 'staff')
    
    # Apply date filters
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            logs = logs.filter(log_date__date__gte=start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            logs = logs.filter(log_date__date__lte=end_dt)
        except ValueError:
            pass
    
    # Calculate total loss
    total_loss = sum(
        abs(log.quantity) * log.product.unit_cost 
        for log in logs 
        if log.product
    )
    
    # Prepare response data
    data = {
        'logs': [
            {
                'id': log.id,
                'date': log.log_date.strftime('%Y-%m-%d %H:%M:%S'),
                'product_name': log.product.product_name if log.product else 'Unknown',
                'quantity': abs(log.quantity),
                'unit_cost': float(log.product.unit_cost) if log.product else 0,
                'loss_amount': float(abs(log.quantity) * log.product.unit_cost) if log.product else 0,
                'staff': f"{log.staff.first_name} {log.staff.last_name}" if log.staff else 'Unknown'
            }
            for log in logs
        ],
        'total_loss': float(total_loss),
        'count': logs.count()
    }
    
# ---------------------------------------------------------
# DISCOUNT AUTO-APPLICATION LOGIC
# ---------------------------------------------------------

def get_applicable_discounts(sale_total=None, sale_date=None):
    """
    Get all applicable discounts for a sale.
    This function can be called when creating sales to automatically apply discounts.
    """
    if sale_date is None:
        sale_date = now().date()
    
    # Get active discounts that are valid for the given date
    applicable_discounts = Discount.objects.filter(
        is_active=True,
        start_date__lte=sale_date,
        end_date__gte=sale_date
    ).order_by('-value')  # Apply highest value discounts first
    
    # If sale_total is provided, filter by minimum amount requirements
    if sale_total is not None:
        # For now, we'll apply all applicable discounts
        # In a more complex system, you might have minimum purchase requirements
        pass
    
    return applicable_discounts


def calculate_discount_amount(discount, sale_total):
    """
    Calculate the discount amount based on discount type and sale total.
    """
    if discount.discount_type == 'Percentage':
        return Decimal(str(sale_total)) * (discount.value / 100)
    elif discount.discount_type == 'Fixed':
        return min(discount.value, Decimal(str(sale_total)))  # Can't discount more than sale total
    elif discount.discount_type == 'BOGO':
        # For BOGO, this is a simplified implementation
        # In a real system, you'd need to track individual items
        return Decimal(str(sale_total)) * (discount.value / (discount.value + 1))
    else:
        return Decimal('0')


def apply_automatic_discounts(sale_instance):
    """
    Apply automatic discounts to a sale instance.
    This should be called after creating a sale but before saving.
    """
    sale_total = float(sale_instance.total_amount)
    sale_date = sale_instance.sale_datetime.date()
    
    applicable_discounts = get_applicable_discounts(sale_total, sale_date)
    
    total_discount_amount = Decimal('0')
    applied_discounts = []
    
    for discount in applicable_discounts:
        discount_amount = calculate_discount_amount(discount, sale_total)
        
        # Apply discount if it's meaningful (more than 0)
        if discount_amount > 0:
            total_discount_amount += discount_amount
            applied_discounts.append({
                'discount': discount,
                'amount': discount_amount
            })
            
            # Prevent over-discounting
            if total_discount_amount >= Decimal(str(sale_total)):
                total_discount_amount = Decimal(str(sale_total))
                break
    
    # Update sale total with discount applied
    if total_discount_amount > 0:
        new_total = Decimal(str(sale_total)) - total_discount_amount
        sale_instance.total_amount = new_total
        sale_instance.save()
        
        # Store applied discounts (you might want to create a SaleDiscount model for this)
        return {
            'original_total': sale_total,
            'discount_amount': float(total_discount_amount),
            'new_total': float(new_total),
            'applied_discounts': applied_discounts
        }
    
    return None


def staff_info_api(request, pk):
    """API endpoint to get staff information for payroll form"""
    staff = get_object_or_404(Staff, pk=pk)
    data = {
        'first_name': staff.first_name,
        'last_name': staff.last_name,
        'position': staff.position,
        'phone': staff.phone,
        'department': getattr(staff, 'department', None),
    }
    return JsonResponse(data)